"""Database client.

PostgreSQL via a threaded connection pool when DATABASE_URL is set;
otherwise an in-memory store guarded by a lock for local/dev use.

Schema is owned by Alembic. This module no longer creates tables at
runtime — run `alembic upgrade head` before serving traffic.
"""

import io
import json
import threading
from contextlib import contextmanager
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Iterator, List, Optional

import bcrypt

from .config import settings
from .errors import APIError
from .logger import get_logger


logger = get_logger(__name__)

_memory_lock = threading.RLock()
_memory_jobs: Dict[str, Dict] = {}
_memory_metrics: Dict[str, Dict] = {}
_memory_users: Dict[str, Dict] = {}

_pool = None
_pool_lock = threading.Lock()


def _get_pool():
    global _pool
    if _pool is not None:
        return _pool
    with _pool_lock:
        if _pool is not None:
            return _pool
        if not settings.database_configured:
            return None
        from psycopg2.pool import ThreadedConnectionPool

        _pool = ThreadedConnectionPool(
            minconn=settings.db_pool_min,
            maxconn=settings.db_pool_max,
            dsn=settings.database_url,
        )
        logger.info("Postgres pool initialized (min=%d max=%d)", settings.db_pool_min, settings.db_pool_max)
        return _pool


@contextmanager
def _conn() -> Iterator[Any]:
    pool = _get_pool()
    if pool is None:
        raise RuntimeError("Database not configured")
    connection = pool.getconn()
    try:
        yield connection
        connection.commit()
    except Exception:
        connection.rollback()
        raise
    finally:
        pool.putconn(connection)


def is_database_available() -> bool:
    return settings.database_configured


# ---------- Jobs ----------


def create_job(
    job_id: str,
    campaign_date: date,
    campaign_name: str,
    product_name: str,
    company: str,
    filename: str,
    file_type: str,
) -> bool:
    if not is_database_available():
        with _memory_lock:
            _memory_jobs[job_id] = {
                "id": job_id,
                "campaign_date": campaign_date,
                "campaign_name": campaign_name,
                "product_name": product_name,
                "company": company,
                "filename": filename,
                "file_type": file_type,
                "status": "processing",
                "created_at": datetime.now(),
                "completed_at": None,
                "error_message": None,
            }
        return True

    try:
        with _conn() as c, c.cursor() as cur:
            cur.execute(
                """
                INSERT INTO jobs (id, campaign_date, campaign_name, product_name,
                                  company, filename, file_type, status)
                VALUES (%s, %s, %s, %s, %s, %s, %s, 'processing')
                """,
                (job_id, campaign_date, campaign_name, product_name, company, filename, file_type),
            )
            return True
    except Exception as exc:  # noqa: BLE001
        logger.error("create_job failed: %s", exc)
        return False


def update_job_status(job_id: str, status: str, error_message: Optional[str] = None) -> bool:
    if not is_database_available():
        with _memory_lock:
            job = _memory_jobs.get(job_id)
            if job is None:
                return False
            job["status"] = status
            if status in {"completed", "failed"}:
                job["completed_at"] = datetime.now()
            if error_message:
                job["error_message"] = error_message
            return True

    try:
        with _conn() as c, c.cursor() as cur:
            if status in {"completed", "failed"}:
                cur.execute(
                    "UPDATE jobs SET status=%s, completed_at=NOW(), error_message=%s WHERE id=%s",
                    (status, error_message, job_id),
                )
            else:
                cur.execute("UPDATE jobs SET status=%s WHERE id=%s", (status, job_id))
            return True
    except Exception as exc:  # noqa: BLE001
        logger.error("update_job_status failed: %s", exc)
        return False


def save_job_metrics(job_id: str, metrics: dict) -> bool:
    if not is_database_available():
        with _memory_lock:
            _memory_metrics[job_id] = {
                "job_id": job_id,
                "total_frames": metrics.get("total_frames", 0),
                "good_frames": metrics.get("good_frames", 0),
                "bad_frames": metrics.get("bad_frames", 0),
                "processing_time_seconds": metrics.get("processing_time_seconds"),
                "metrics_json": metrics.get("metrics_json", {}),
                "created_at": datetime.now(),
            }
        return True

    try:
        with _conn() as c, c.cursor() as cur:
            cur.execute(
                """
                INSERT INTO job_metrics (job_id, total_frames, good_frames, bad_frames,
                                         processing_time_seconds, metrics_json)
                VALUES (%s, %s, %s, %s, %s, %s::jsonb)
                ON CONFLICT (job_id) DO UPDATE SET
                    total_frames = EXCLUDED.total_frames,
                    good_frames = EXCLUDED.good_frames,
                    bad_frames = EXCLUDED.bad_frames,
                    processing_time_seconds = EXCLUDED.processing_time_seconds,
                    metrics_json = EXCLUDED.metrics_json
                """,
                (
                    job_id,
                    metrics.get("total_frames", 0),
                    metrics.get("good_frames", 0),
                    metrics.get("bad_frames", 0),
                    metrics.get("processing_time_seconds"),
                    json.dumps(metrics.get("metrics_json", {})),
                ),
            )
            return True
    except Exception as exc:  # noqa: BLE001
        logger.error("save_job_metrics failed: %s", exc)
        return False


def get_job_by_id(job_id: str) -> Optional[Dict[str, Any]]:
    if not is_database_available():
        with _memory_lock:
            job = _memory_jobs.get(job_id)
            if not job:
                return None
            merged = job.copy()
            if job_id in _memory_metrics:
                merged.update(_memory_metrics[job_id])
            return merged

    try:
        from psycopg2.extras import RealDictCursor

        with _conn() as c, c.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                """
                SELECT j.*, m.total_frames, m.good_frames, m.bad_frames,
                       m.processing_time_seconds, m.metrics_json
                FROM jobs j LEFT JOIN job_metrics m ON j.id = m.job_id
                WHERE j.id = %s
                """,
                (job_id,),
            )
            row = cur.fetchone()
            return dict(row) if row else None
    except Exception as exc:  # noqa: BLE001
        logger.error("get_job_by_id failed: %s", exc)
        return None


def get_all_jobs(limit: int = 100, offset: int = 0, status_filter: Optional[str] = None) -> List[Dict[str, Any]]:
    if not is_database_available():
        with _memory_lock:
            jobs = []
            for job_id, job in _memory_jobs.items():
                if status_filter and job.get("status") != status_filter:
                    continue
                merged = job.copy()
                if job_id in _memory_metrics:
                    merged.update(_memory_metrics[job_id])
                jobs.append(merged)
        jobs.sort(key=lambda j: j.get("created_at", datetime.min), reverse=True)
        return jobs[offset : offset + limit]

    try:
        from psycopg2.extras import RealDictCursor

        with _conn() as c, c.cursor(cursor_factory=RealDictCursor) as cur:
            base = """
                SELECT j.*, m.total_frames, m.good_frames, m.bad_frames,
                       m.processing_time_seconds, m.metrics_json
                FROM jobs j LEFT JOIN job_metrics m ON j.id = m.job_id
            """
            if status_filter:
                cur.execute(
                    base + " WHERE j.status = %s ORDER BY j.created_at DESC LIMIT %s OFFSET %s",
                    (status_filter, limit, offset),
                )
            else:
                cur.execute(
                    base + " ORDER BY j.created_at DESC LIMIT %s OFFSET %s",
                    (limit, offset),
                )
            return [dict(row) for row in cur.fetchall()]
    except Exception as exc:  # noqa: BLE001
        logger.error("get_all_jobs failed: %s", exc)
        return []


# ---------- Excel export ----------

_METRIC_KEYS = (
    "views",
    "accounts_reached",
    "followers",
    "non_followers",
    "interactions",
    "likes",
    "replies",
    "shares",
    "profile_visits",
    "follows",
    "links_clicks",
    "sticker_taps",
    "navigation",
    "forward",
    "next_story",
    "back",
    "exited",
    "profile_activity",
    "external_link_taps",
)


def export_to_excel(output_path: Optional[Path] = None) -> Optional[bytes]:
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError as exc:
        raise APIError(500, "openpyxl not installed") from exc

    jobs = get_all_jobs(limit=10000)
    if not jobs:
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Campaign Results"

    headers = [
        "Job ID", "Campaign Date", "Campaign Name", "Product Name", "Company",
        "Filename", "File Type", "Status", "Created At", "Completed At",
        "Total Frames", "Good Frames", "Bad Frames", "Processing Time (s)",
    ] + [k.replace("_", " ").title() for k in _METRIC_KEYS]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    alt_fill = PatternFill(start_color="E9EBF5", end_color="E9EBF5", fill_type="solid")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for row_idx, job in enumerate(jobs, 2):
        metrics_json = job.get("metrics_json") or {}
        summary = metrics_json.get("summary", {}) if isinstance(metrics_json, dict) else {}

        row = [
            str(job.get("id", "")),
            job["campaign_date"].isoformat() if job.get("campaign_date") else "",
            job.get("campaign_name", ""),
            job.get("product_name", ""),
            job.get("company", ""),
            job.get("filename", ""),
            job.get("file_type", ""),
            job.get("status", ""),
            job["created_at"].isoformat() if job.get("created_at") else "",
            job["completed_at"].isoformat() if job.get("completed_at") else "",
            job.get("total_frames", 0),
            job.get("good_frames", 0),
            job.get("bad_frames", 0),
            job.get("processing_time_seconds", 0),
        ] + [summary.get(k, "") for k in _METRIC_KEYS]

        for col, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = border
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    widths = [36, 12, 20, 20, 20, 25, 10, 12, 20, 20, 12, 12, 12, 15] + [14] * len(_METRIC_KEYS)
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    ws.freeze_panes = "A2"

    if output_path:
        wb.save(output_path)
        return None

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------- Users / auth ----------


def _hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")


def _verify_password(password: str, stored_hash: str) -> bool:
    if not stored_hash:
        return False
    if stored_hash.startswith("$2"):
        try:
            return bcrypt.checkpw(password.encode("utf-8"), stored_hash.encode("utf-8"))
        except ValueError:
            return False
    # Legacy SHA-256 hash with hardcoded salt — verify and rehash on next login.
    import hashlib

    legacy = hashlib.sha256(f"instagram_analyzer_salt_2024{password}".encode()).hexdigest()
    return legacy == stored_hash


def _is_legacy_hash(stored_hash: str) -> bool:
    return bool(stored_hash) and not stored_hash.startswith("$2")


def create_user(username: str, password: str) -> bool:
    password_hash = _hash_password(password)

    if not is_database_available():
        with _memory_lock:
            if username in _memory_users:
                return False
            _memory_users[username] = {
                "username": username,
                "password_hash": password_hash,
                "created_at": datetime.now(),
            }
            return True

    try:
        with _conn() as c, c.cursor() as cur:
            cur.execute(
                "INSERT INTO users (username, password_hash) VALUES (%s, %s)",
                (username, password_hash),
            )
            return True
    except Exception as exc:  # noqa: BLE001
        logger.error("create_user failed: %s", exc)
        return False


def _update_user_password(username: str, password: str) -> None:
    new_hash = _hash_password(password)
    if not is_database_available():
        with _memory_lock:
            user = _memory_users.get(username)
            if user:
                user["password_hash"] = new_hash
        return
    try:
        with _conn() as c, c.cursor() as cur:
            cur.execute(
                "UPDATE users SET password_hash=%s WHERE username=%s",
                (new_hash, username),
            )
    except Exception as exc:  # noqa: BLE001
        logger.error("rehash failed for %s: %s", username, exc)


def verify_user(username: str, password: str) -> bool:
    if not is_database_available():
        with _memory_lock:
            user = _memory_users.get(username)
        if not user:
            return False
        stored = user["password_hash"]
    else:
        try:
            with _conn() as c, c.cursor() as cur:
                cur.execute("SELECT password_hash FROM users WHERE username=%s", (username,))
                row = cur.fetchone()
                if not row:
                    return False
                stored = row[0]
        except Exception as exc:  # noqa: BLE001
            logger.error("verify_user failed: %s", exc)
            return False

    if not _verify_password(password, stored):
        return False

    if _is_legacy_hash(stored):
        logger.info("Upgrading legacy password hash for %s", username)
        _update_user_password(username, password)
    return True


def get_user_count() -> int:
    if not is_database_available():
        with _memory_lock:
            return len(_memory_users)
    try:
        with _conn() as c, c.cursor() as cur:
            cur.execute("SELECT COUNT(*) FROM users")
            row = cur.fetchone()
            return int(row[0]) if row else 0
    except Exception as exc:  # noqa: BLE001
        logger.error("get_user_count failed: %s", exc)
        return 0
